
"""
A python program that ranks images based on the glicko rating system(https://en.wikipedia.org/wiki/Glicko_rating_system) and user input

"""

import argparse
import glob
import json
import os
import sys
import math
import openpyxl
from openpyxl import Workbook, load_workbook
import shutil
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.image as mpimg
import exifread
from imagededup.methods import CNN



class Photo:

    LEFT = 0
    RIGHT = 1
    DRAW = 0,5

    c = math.sqrt((pow(350,2) - pow(50,2))/20)

    def __init__(self, filename, score = 1400.0, wins = 0, matches = 0, kvalue = 350):

        if not os.path.isfile(filename):
            raise ValueError("No file named {filename}")

        self._filename = filename
        self._score = score
        self._wins = wins
        self._matches = matches
        self._kvalue = kvalue

        self._read_and_downsample()


    def data(self):
        return self._data


    def filename(self):
        return self._filename


    def matches(self):
        return self._matches

    def kvalue(self):
        return self._kvalue



    def score(self, k = None, s = None, is_winner = None):

        if s is None:
            return self._score

        assert is_winner is not None

        self._score = s

        self._matches += 1

        self._kvalue = k

        if is_winner:
            self._wins += 1


    def win_percentage(self):
        if self._matches == 0:
            return 0.0
        else:
            return 100.0 * float(self._wins) / float(self._matches)


    def __eq__(self, rhs):
        return self._filename == rhs._filename


    def to_dict(self):

        return {
            'filename' : self._filename,
            'score' : self._score,
            'matches' : self._matches,
            'wins' : self._wins,
            'kvalue' : self._kvalue,
        }


    def _read_and_downsample(self):
        """
        Reads the image, performs rotation, and downsamples.
        """

        #----------------------------------------------------------------------
        # read image

        f = self._filename

        data = mpimg.imread(f)

        #----------------------------------------------------------------------
        # downsample

        # the point of downsampling is so the images can be redrawn by the
        # display as fast as possible, this is so one can iterate though the
        # image set as quickly as possible.  No one want's to wait around for
        # the fat images to be loaded over and over.

        # dump downsample, just discard columns-n-rows

        M, N = data.shape[0:2]

        MN = max([M,N])

        step = int(MN / 800)
        if step == 0: step = 1

        data = data[ 0:M:step, 0:N:step, :]

        #----------------------------------------------------------------------
        # rotate

        # read orientation with exifread

        with open(f, 'rb') as fd:
            tags = exifread.process_file(fd)

        r = 'Horizontal (normal)'

        try:
            r = str(tags['Image Orientation'])
        except:
            pass

        # rotate as necessary

        if r == 'Horizontal (normal)':
            pass

        elif r == 'Rotated 90 CW':

            data = np.rot90(data, 3)

        elif r == 'Rotated 90 CCW':

            data = np.rot90(data, 1)

        elif r == 'Rotated 180':

            data = np.rot90(data, 2)

        else:
            raise RuntimeError('Unhandled rotation "%s"' % r)

        self._data = data


class Display(object):
    """
    Given two photos, displays them with Matplotlib and provides a graphical
    means of choosing the better photo.

    Click on the select button to pick the better photo.

    ~OR~

    Press the left or right arrow key to pick the better photo.

    """


    def __init__(self, f1, f2, title = None, figsize = None, duplicates = False):

        self._choice = None
        assert isinstance(f1, Photo)
        assert isinstance(f2, Photo)

        if figsize is None:
            figsize = [20,12]

        fig = plt.figure(figsize=figsize)
        plt.get_current_fig_manager().window.wm_geometry("+1+5")

        h = 10

        ax11 = plt.subplot2grid((h,2), (0,0), rowspan = h - 1)
        ax12 = plt.subplot2grid((h,2), (0,1), rowspan = h - 1)

        ax21 = plt.subplot2grid((h,6), (h - 1, 1))
        ax22 = plt.subplot2grid((h,6), (h - 1, 4))
        ax23 = plt.subplot2grid((h,6), (h - 1, 2), colspan = 2)
        if duplicates == False:
            kwargs1 = dict(s = 'Select', ha = 'center', va = 'center', fontsize=20)
            kwargs2 = dict(s = 'Draw', ha = 'center', va = 'center', fontsize=20)
        elif duplicates == True:
            kwargs1 = dict(s = 'Delete', ha = 'center', va = 'center', fontsize=20)
            kwargs2 = dict(s = 'Keep both', ha = 'center', va = 'center', fontsize=20)

        ax21.text(0.5, 0.5, **kwargs1)
        ax22.text(0.5, 0.5, **kwargs1)
        ax23.text(0.5, 0.5, **kwargs2)

        self._fig = fig
        self._ax_select_left = ax21
        self._ax_select_right = ax22
        self._ax_select_draw = ax23

        fig.subplots_adjust(
            left = 0.02,
            bottom = 0.02,
            right = 0.98,
            top = 0.98,
            wspace = 0.05,
            hspace = 0,
        )

        ax11.imshow(f1.data())
        ax12.imshow(f2.data())

        for ax in [ax11, ax12, ax21, ax22, ax23]:
            ax.set_xticklabels([])
            ax.set_yticklabels([])

            ax.set_xticks([])
            ax.set_yticks([])

        self._attach_callbacks()

        if title:
            fig.suptitle(title, fontsize=20)

        plt.show()


    def _on_click(self, event):

        if event.inaxes == self._ax_select_left:
            self._choice = Photo.LEFT
            plt.close(self._fig)

        elif event.inaxes == self._ax_select_right:
            self._choice = Photo.RIGHT
            plt.close(self._fig)

        elif event.inaxes == self._ax_select_draw:
            self._choice = Photo.DRAW
            plt.close(self._fig)

    def _on_key_press(self, event):

        if event.key == 'left':
            self._choice = Photo.LEFT
            plt.close(self._fig)

        elif event.key == 'right':
            self._choice = Photo.RIGHT
            plt.close(self._fig)


    def _attach_callbacks(self):
        self._fig.canvas.mpl_connect('button_press_event', self._on_click)
        self._fig.canvas.mpl_connect('key_press_event', self._on_key_press)


class EloTable:


    def __init__(self, max_increase = 32.0):
        self._K = max_increase
        self._photos = {}
        self._shuffled_keys = []


    def add_photo(self, filename_or_photo, filepath = None):


        wb = load_workbook('ranked.xlsx')
        ws = wb.active
        g = 2

        if isinstance(filename_or_photo, str):

            filename = filename_or_photo

            if filename not in self._photos:
                self._photos[filename] = Photo(filename)
                if filepath != None:
                    while ws.cell(column = 1, row = g).value != None:
                        g += 1;
                    c1 = ws.cell(column = 1, row = g)
                    c1.value = filename
                    wb.save(filepath + '/ranked.xlsx')
        elif isinstance(filename_or_photo, Photo):

            photo = filename_or_photo

            if photo.filename() not in self._photos:
                self._photos[photo.filename()] = photo
                if filepath != None:
                    while ws.cell(column = 1, row = g).value != None:
                        g += 1;
                    c1 = ws.cell(column = 1, row = g)
                    c1.value = photo.filename()
                    wb.save(filepath + '/ranked.xlsx')

        # Convert the dictionary into a list and then sort by score.
    def get_ranked_list(self):



        ranked_list = self._photos.values()

        ranked_list = sorted(
            ranked_list,
            key = lambda record : record.score(),
            reverse = True)

        return ranked_list

    def rank_photos(self, n_iterations, figsize):
        """
        Displays two photos using the command "gnome-open".  Then asks which
        photo is better.
        """

        n_photos = len(self._photos)

        keys = list(self._photos.keys())

        for i in range(n_iterations):

            np.random.shuffle(keys)

            n_matchups = n_photos / 2

            for j in range(0, n_photos - 1, 2):

                match_up = j / 2

                title = 'Round %d / %d, Match Up %d / %d' % (
                    i + 1, n_iterations,
                    match_up + 1,
                    n_matchups)

                photo_a = self._photos[keys[j]]
                photo_b = self._photos[keys[j+1]]

                d = Display(photo_a, photo_b, title, figsize,duplicates=False)

                if d._choice == Photo.LEFT:
                    self.__score_result(photo_a, photo_b, draw = False)
                elif d._choice == Photo.RIGHT:
                    self.__score_result(photo_b, photo_a, draw = False)
                elif d._choice == Photo.DRAW:
                    self.__score_result(photo_b, photo_a, draw = True)
                else:
                    raise RuntimeError("oops, found a bug!")



    def __score_result(self, winning_photo, loosing_photo, draw):

        # Current ratings
        R_a = winning_photo.score()
        R_b = loosing_photo.score()

        # Current rating deviation
        k_a = winning_photo.kvalue()
        k_b = loosing_photo.kvalue()

        q = 0.00575646273

        g_a = 1 / math.sqrt(1 + 3 * pow(q , 2) * pow(k_a, 2) / pow(math.pi, 2))
        g_b = 1 / math.sqrt(1 + 3 * pow(q , 2) * pow(k_b, 2) / pow(math.pi, 2))

        # Expectation
        E_a = 1.0 / (1.0 + 10.0 ** (-1 * g_a * (R_a - R_b) / 400.0))
        E_b = 1.0 / (1.0 + 10.0 ** (-1 * g_b * (R_b - R_a) / 400.0))

        d_a = math.sqrt(1/(pow(q, 2) * pow(g_a, 2) * E_a * (1 - E_a)))
        d_b = math.sqrt(1/(pow(q, 2) * pow(g_b, 2) * E_b * (1 - E_b)))

        # New rating deviation
        k_a = 1 / math.sqrt((1/(pow(k_a, 2)) + (1/pow(d_a, 2))))
        k_b = 1 / math.sqrt((1/(pow(k_b, 2)) + (1/pow(d_b, 2))))

        # New ratings
        if draw == True:
            R_a = R_a + ((q * pow(k_a,2) * g_a) * (0.5 - E_a))
            R_b = R_b + ((q * pow(k_b,2) * g_b) * (0.5 - E_b))

            winning_photo.score(k_a,R_a, False)
            loosing_photo.score(k_b,R_b, False)

        else:
            R_a = R_a + ((q * pow(k_a,2) * g_a) * (1.0 - E_a))
            R_b = R_b + ((q * pow(k_b,2) * g_b) * (0.0 - E_b))
            winning_photo.score(k_a,R_a, True)
            loosing_photo.score(k_b,R_b, False)




    def to_dict(self):

        rl = self.get_ranked_list()

        rl = [x.to_dict() for x in rl]

        return {'photos' : rl}

def find_duplicates1(figsize, directory):
    if not os.path.exists('Duplicates'):
        os.makedirs('Duplicates')
    print("Finding duplicates...")
    method_object = CNN()
    duplicates = method_object.find_duplicates(image_dir=directory)

    for key, value in duplicates.items():
        for i, item in enumerate(value):
            if isinstance(item[1], np.float32):
                # Create a new tuple with the modified value
                new_item = (item[0], float(item[1]))
                # Replace the original tuple with the new one
                value[i] = new_item

    for key, value in list(duplicates.items()):
        # Check if the value is not an empty list
        if not value:
            # Add the key and value to the list
            del duplicates[key]

    # Open a file for writing
    with open('duplicates.json', 'w') as outfile:
        # Write the dictionary to the file in JSON format
        json.dump(duplicates, outfile)

    already_shown = []
    for key, value_list in duplicates.items():
        if key not in already_shown:
            value_list.insert(0,key)
            for index, value in enumerate(value_list):
                if index + 1 < len(value_list):
                # Skip over any values that have already been moved to the Duplicates directory
                    if os.path.exists(value) and os.path.exists(value_list[index + 1]):
                        photo_a = Photo(value_list[index])
                        photo_b = Photo(value_list[index + 1])
                        d = Display(photo_a, photo_b,title='Duplicates',figsize=figsize,duplicates=True)
                        if d._choice == Photo.LEFT:
                            shutil.move(directory + '\\' + photo_a._filename,directory + '\\Duplicates')
                        elif d._choice == Photo.RIGHT:
                            shutil.move(directory + '\\' + photo_b._filename,directory + '\\Duplicates')
                        elif d._choice == Photo.DRAW:
                            pass
                        else:
                            raise RuntimeError("oops, found a bug!")
        already_shown.append(key)
        already_shown.extend(value_list)
    print('done')


def rank(rounds,figsize,directory,table):


    #--------------------------------------------------------------------------
    # Rank the photos!

    table.rank_photos(rounds, figsize)

    #--------------------------------------------------------------------------
    # save the table

    with open(directory + '\\ranking_table.json', 'w') as fd:

        d = table.to_dict()

        jstr = json.dumps(d, indent = 4, separators=(',', ' : '))

        fd.write(jstr)

    #--------------------------------------------------------------------------
    # dump ranked list to disk
    print('123')
    with open(directory + '\\ranked.txt', 'w') as fd:

        ranked_list = table.get_ranked_list()

        heading_fmt = "%4d    %4.0f    %7d    %7.2f    %s\n"

        heading = "Rank    Score    Matches    Win %    Filename\n"

        fd.write(heading)

        for i, photo in enumerate(ranked_list):

            line = heading_fmt %(
                i + 1,
                photo.score(),
                photo.matches(),
                photo.win_percentage(),
                photo.filename())



            fd.write(line)
    print('123')
    #--------------------------------------------------------------------------
    # write xsl table
    wb = load_workbook(directory + '/ranked.xlsx')
    ws = wb.active
    a = 2

    for g, photo in enumerate(ranked_list):
        g +=1

    while ws.cell(row = 1,column = a).value != None:
        a += 1
    c1 = ws.cell(row = 1, column = a)
    c1.value = a - 1
    for z, photo in enumerate(ranked_list):
        b = 2
        while(True):
            if ws.cell(row = b,column = 1).value == photo.filename():
                c2 = ws.cell(row = b, column = a)
                c2.value = photo.score()
                break;
            b += 1

    wb.save(directory + '/ranked.xlsx')
    #--------------------------------------------------------------------------
    # dump ranked list to screen

    print("Final Ranking:")

    with open(directory + '\\ranked.txt', 'r') as fd:
        text = fd.read()

    print(text)

def main():
    description = """\
    Uses the Elo ranking algorithm to sort your images by rank.  The program globs
    for .jpg images to present to you in random order, then you select the better
    photo.  After n-rounds, the results are reported.

    Click on the "Select" button or press the LEFT or RIGHT arrow to pick the
    better photo.

    """
    parser = argparse.ArgumentParser(description = description)

    parser.add_argument(
        '-r', '--rank',
        action='store_true',
        help='rank images'
    )

    parser.add_argument(
        '-d', '--duplicates',
        action='store_true',
        help='display duplicates'
    )

    parser.add_argument(
        "-ro",
        "--n-rounds",
        type = int,
        default = 3,
        help = "Specifies the number of rounds to pass through the photo set (3)"
    )

    parser.add_argument(
        "-fi",
        "--figsize",
        nargs=2,
        type=int,
        default=[20, 12],
        help="Specifies width and height of the Matplotlib figsize (20, 12)"
    )




    parser.add_argument(
        "photo_dir",
        help = "The photo directory to scan for .jpg images"
    )

    args = parser.parse_args()

    assert os.path.isdir(args.photo_dir)

    os.chdir(args.photo_dir)

    ranking_table_json = 'ranking_table.json'
    ranked_txt         = 'ranked.txt'
    if not os.path.isfile('ranked.xlsx'):
        wb = openpyxl.Workbook()
        wb.save(args.photo_dir + '/ranked.xlsx')



    # Create the ranking table and add photos to it.

    table = EloTable()

    #--------------------------------------------------------------------------
    # Read in table .json if present

    sys.stdout.write("Reading in photos and downsampling ...")
    sys.stdout.flush()

    if os.path.isfile(ranking_table_json):
        with open(ranking_table_json, 'r') as fd:
            d = json.load(fd)

        filtered_photos = []
        for i, photo in enumerate(d['photos']):
            filename = photo["filename"]
            filepath = os.path.join(args.photo_dir, filename)
            if not os.path.isfile(filepath):
                print('deleting ' + filename + '...')
                filtered_photos.append(photo)
        for photo in filtered_photos:
            d['photos'].remove(photo)

# Write the modified data back to the JSON file
        with open(ranking_table_json, 'w') as f:
            json.dump(d, f)


        # read photos and add to table
    if os.path.isfile(ranking_table_json):
        with open(ranking_table_json, 'r') as fd:
            d = json.load(fd)

        for p in d['photos']:

            photo = Photo(**p)

            table.add_photo(photo)

    #--------------------------------------------------------------------------
    # glob for files, to include newly added files

    filelist = glob.glob('*.jpg') + glob.glob('*.jpeg')

    for f in filelist:
        table.add_photo(f, args.photo_dir)



    print("Done !")

    if args.duplicates:
        find_duplicates1(args.figsize,args.photo_dir)
        if os.path.isfile(ranking_table_json):
            with open(ranking_table_json, 'r') as fd:
                d = json.load(fd)

            filtered_photos = []
            for i, photo in enumerate(d['photos']):
                filename = photo["filename"]
                filepath = os.path.join(args.photo_dir, filename)
                if not os.path.isfile(filepath):
                    print('deleting ' + filename + '...')
                    filtered_photos.append(photo)
            for photo in filtered_photos:
                d['photos'].remove(photo)

    # Write the modified data back to the JSON file
            with open(ranking_table_json, 'w') as f:
                json.dump(d, f)
    if args.rank:
        rank(args.n_rounds,args.figsize,args.photo_dir,table)

        # do something else




if __name__ == "__main__": main()
